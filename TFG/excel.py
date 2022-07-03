from openpyxl import load_workbook
import math
from db_functions import *
from pdf_read import *

'''
A2_Estudio_Donante_Sano
A3_Estudio_Pre_Transplante
B1_Contaje_Control_Pre_Aféresis
B2_Contaje_1er_Tubo
B3_Contaje_2o_Tubo
B4_Contaje_Controles_Intermedios
B51_Contaje_Producto_Final
B52_Cultivo_Micro_Final
B53_Etiqueta_Aféresis
B54_Etiquetas_Congelación
C1_Etiqueta_Producto_Inicial
C21_Contaje_Post_Sepax
C22_Cultivo_Microbiologico_Post_Sepax
C31_Contaje_Frac_Deplec
C32_Cultivo_Microbiologico_Frac_Deplec
C41_Contaje_Frac_Pos
C42_Cultivo_Microbiologico_Frac_Pos
C5_Etiqueta_Infusión
C11_Contaje_Control_NUNC
C12_Cultivo_Microbiológico_NUNC
C2_Contajes_Bolsas_Descong_sin_lavar
C31_Contajes_Bolsas_Descong_lavadas
C32_Etiquetas_Descongelación
'''

def informe_aferesis(donation_code, file, dst):
    src = "./plantillas/Plantilla_Aferesis.xlsx"

    # Leer Datos procesamiento

    wb = load_workbook(src)
    sheets = wb.sheetnames

    sheet = wb[sheets[0]]

    r, c = format_cell("B2")
    sheet.cell(row = r, column = c).value = donation_code[-6:]
    r, c = format_cell("M2")
    sheet.cell(row = r, column = c).value = get_info("A3_Estudio_Pre_Transplante", "Peso", donation_code)
    r, c = format_cell("B3")
    sheet.cell(row = r, column = c).value = get_info("B51_Contaje_Producto_Final", "Fecha", donation_code)
    r, c = format_cell("B4")
    sheet.cell(row = r, column = c).value = get_info("A3_Estudio_Pre_Transplante", "Tipo_procesador_celular", donation_code)
    r, c = format_cell("B14")
    sheet.cell(row = r, column = c).value = get_info("B51_Contaje_Producto_Final", "Linfocitos_porc", donation_code)
    r, c = format_cell("B16")
    sheet.cell(row = r, column = c).value = get_info("B51_Contaje_Producto_Final", "Monocitos_porc", donation_code)
    r, c = format_cell("B19")
    sheet.cell(row = r, column = c).value = get_info("B51_Contaje_Producto_Final", "CD34pos_porc", donation_code)
    r, c = format_cell("B23")
    sheet.cell(row = r, column = c).value = get_info("B51_Contaje_Producto_Final", "Leucosviables_porc", donation_code)

    r, c = format_cell("B37")
    sheet.cell(row = r, column = c).value = get_info("B2_Contaje_1er_Tubo", "Hemoglobina", donation_code)
    r, c = format_cell("B38")
    sheet.cell(row = r, column = c).value = get_info("B3_Contaje_2o_Tubo", "Hemoglobina", donation_code)
    r, c = format_cell("B41")
    sheet.cell(row = r, column = c).value = str(math.trunc(float(get_info("B2_Contaje_1er_Tubo", "Plaquetas", donation_code)))*1000)
    r, c = format_cell("B42")
    sheet.cell(row = r, column = c).value = str(math.trunc(float(get_info("B3_Contaje_2o_Tubo", "Plaquetas", donation_code)))*1000)

    sheet = wb[sheets[1]]
    
    nhc, nombre, apellidos, fecha_nac = data_type2(file)

    r, c = format_cell("B2")
    sheet.cell(row = r, column = c).value = nombre
    r, c = format_cell("B3")
    sheet.cell(row = r, column = c).value = apellidos
    r, c = format_cell("E2")
    sheet.cell(row = r, column = c).value = nhc
    r, c = format_cell("E3")
    sheet.cell(row = r, column = c).value = fecha_nac
    r, c = format_cell("F3")
    sheet.cell(row = r, column = c).value = get_info("A3_Estudio_Pre_Transplante", "Peso", donation_code)

    wb.save(dst)

def informe_infusion(donation_code, file, dst):
    src = "./plantillas/Plantilla_Infusion.xlsx"

    wb = load_workbook(src)
    sheets = wb.sheetnames

    sheet = wb[sheets[0]]

    numbers = ["5", "7", "9", "16", "20", "23", "30", "32", "34", "36"]
    letters = ["B", "E", "F", "G"]
    tables = ["B51_Contaje_Producto_Final", "C21_Contaje_Post_Sepax", "C41_Contaje_Frac_Pos", "C31_Contaje_Frac_Deplec"]
    rows = ["Neutrofilos_porc", "Linfocitos_porc", "Monocitos_porc", "CD34pos_porc", "CD3pos_porc", "Leucosviables_porc", "CD3posCD8pos_porc", "CD3posCD4pos_porc", "CD56posCD3neg_porc", "CD20pos_porc"]
    
    r, c = format_cell("I2")
    sheet.cell(row = r, column = c).value = get_info("A3_Estudio_Pre_Transplante", "Peso", donation_code)

    for i in range(len(numbers)):
        for j in range(len(letters)):
            cell = letters[j]+numbers[i]
            r, c = format_cell(cell)
            sheet.cell(row = r, column = c).value = get_info(tables[j], rows[i], donation_code)

    r, c = format_cell("B39")
    sheet.cell(row = r, column = c).value = get_info("B2_Contaje_1er_Tubo", "Hemoglobina", donation_code)
    r, c = format_cell("B40")
    sheet.cell(row = r, column = c).value = get_info("B3_Contaje_2o_Tubo", "Hemoglobina", donation_code)
    r, c = format_cell("B43")
    sheet.cell(row = r, column = c).value = str(math.trunc(float(get_info("B2_Contaje_1er_Tubo", "Plaquetas", donation_code)))*1000)
    r, c = format_cell("B44")
    sheet.cell(row = r, column = c).value = str(math.trunc(float(get_info("B3_Contaje_2o_Tubo", "Plaquetas", donation_code)))*1000)

    sheet = wb[sheets[1]]

    r, c = format_cell("B2")
    sheet.cell(row = r, column = c).value = get_info("C42_Cultivo_Microbiologico_Frac_Pos", "Fecha", donation_code)

    sheet = wb[sheets[3]]

    nhc_sano, sano_nombre, sano_apellidos, nhc_receptor, receptor_nombre, receptor_apellidos = data_type1(file)

    r, c = format_cell("B3")
    sheet.cell(row = r, column = c).value = receptor_nombre
    r, c = format_cell("B4")
    sheet.cell(row = r, column = c).value = receptor_apellidos
    r, c = format_cell("E3")
    sheet.cell(row = r, column = c).value = nhc_receptor
    r, c = format_cell("E4")
    sheet.cell(row = r, column = c).value = get_info("A3_Estudio_Pre_Transplante", "Peso", donation_code)
    r, c = format_cell("B5")
    sheet.cell(row = r, column = c).value = sano_nombre
    r, c = format_cell("B6")
    sheet.cell(row = r, column = c).value = sano_apellidos
    r, c = format_cell("E5")
    sheet.cell(row = r, column = c).value = nhc_sano
    r, c = format_cell("E6")
    sheet.cell(row = r, column = c).value = get_info("A2_Estudio_Donante_Sano", "Peso", donation_code)

    wb.save(dst)

def format_cell(cell):
    split = re.split('(\d+)', cell)
    del split[-1]
    c = split[0]
    r = int(split[1])
    c = ord(c) -64
    return r, c